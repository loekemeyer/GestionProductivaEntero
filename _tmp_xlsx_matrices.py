from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

data = [
    (1, "Corte Cuchilla Pelador", "X1"),
    (3, "Corte y Estampado Mango pela", "M10"),
    (6, "Corte y Estampa Varilla", "HF7"),
    (7, "Corte Cuchilla Abrelatas", "IF12"),
    (9, "Doblado Cuchilla", "IF14"),
    (10, "Armado Varilla C/Cuchilla", "HF11"),
    (11, "Corte Mgo Plano Mariposa", "IF4"),
    (12, "Doblado Mango Plano", "IF6"),
    (13, "Corte Mariposa", "G15"),
    (14, "Corte Engranaje chico", "W7"),
    (15, "Corte Arandela fina 502", "W9"),
    (19, "Recortado Manija", "HF1"),
    (20, "Corte Engranaje", "W2"),
    (21, "Corte Arandela buje 501", "W1"),
    (22, "Corte Arandela fina 501", "W3"),
    (27, "Corte Cuerpo Uña Pie", "IF1, JF13"),
    (28, "Corte Cuerpo Uña", "JF2"),
    (29, "Corte Uña", "LF13"),
    (33, "Estampado Cuerpo 3 en 1", "JF10, JF12"),
    (39, "Cerrado Cuerpo Sacacorcho", "KF2, KF5"),
    (40, "Corte Sacatapita", "KF7"),
    (41, "Estampado Sacatapita", "KF8"),
    (44, "Corte Cuchufli", "LF16"),
    (45, "Corte Cabezal", "JF14"),
    (46, "Doblado Cabezal", "JF15"),
    (47, "Corte Resorte Ø 2.75 mm", "IF3"),
    (69, "Doblado Resorte", "IF2"),
    (71, "Corte Arandela Grande", "KF11"),
    (72, "Corte Arandela Chica", "KF9"),
    (74, "Estampado Rompenuez Cerrada", "G7"),
    (77, "Aplastado Punta Rompenuez", "G6, G8"),
    (79, "Corte Destapacorona", "G1, JF8"),
    (81, "Doblado Destapacorona", "G2, JF7"),
    (113, "Remachado pisa papas Inox", "M1"),
    (116, "Corte de Aleta", "LF11, LF12"),
    (127, "Estampado Pza Grande", "Z2B, Z3B"),
    (134, "Remachado pinza chica/Gde", "N7, N8"),
    (137, "Cortar arandela Batidor mini", "LLF7B"),
    (151, "Remachado Sacafuente", "Z36"),
    (155, "Estampado Pza Chica", "Z4"),
    (169, "Doblado vastago pala canelon", "LF8"),
    (173, "Doblado Varilla Curva", "HF6"),
    (174, "Armado Varilla Curva C/Cuchilla", "HF15"),
    (182, "Estampado flecha ahuecapapa/ahuecafruta", "N3"),
    (183, "Soldar Ahuecapapa/Ahuecafruta", "N1, N2"),
    (205, "Ajuste tochos", "JF1"),
    (219, "Hacer Virola", "D13B"),
    (221, "Estampado Aleta", "LF10, LF9"),
    (345, "Corte Vastago Pelapapa Laser", "LLF7A"),
    (346, "Corte Vastago Corta Pizza Grande", "KF14"),
    (349, "Corte Disco Pisapapa", "LLF4"),
    (354, "Corte Vast Corta Pizza Chico", "LF15"),
    (356, "Corte Mango Plano Manija", "G11, IF10"),
    (357, "Corte Eses", "LLF1"),
    (358, "Corte Disco Corta Pizza Chico", "LLF3"),
    (359, "Estampado Corta Raviol", "LLF2"),
    (360, "Corte Ahueca", "N4, N5"),
    (364, "Corte Pieza Chica SacaFuente Pizzero", "Z6"),
    (366, "Corte Super Mariposita", "HF16"),
    (367, "Corte Puntera Pinza Grande", "N6"),
    (368, "Doblado Sacfuente", "Z5"),
]

wb = Workbook()
ws = wb.active
ws.title = "Matrices que suman SC"

headers = ["N° Matriz", "Nombre", "SC que Aumenta"]
ws.append(headers)

for row in data:
    ws.append(row)

# Estilos
font_arial = Font(name="Arial", size=11)
header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill("solid", start_color="305496")
center = Alignment(horizontal="center", vertical="center")
left = Alignment(horizontal="left", vertical="center")
thin = Side(border_style="thin", color="B4B4B4")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Header
for col in range(1, 4):
    c = ws.cell(row=1, column=col)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = border

# Body
for r in range(2, len(data) + 2):
    for col in range(1, 4):
        c = ws.cell(row=r, column=col)
        c.font = font_arial
        c.border = border
        c.alignment = center if col != 2 else left

# Anchos
ws.column_dimensions["A"].width = 12
ws.column_dimensions["B"].width = 42
ws.column_dimensions["C"].width = 22

# Freeze + filtro
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:C{len(data)+1}"

wb.save(r"Z:\AA IT\Gestion Productiva (4)\Matrices_Suman_SC.xlsx")
print("OK", len(data), "filas")
